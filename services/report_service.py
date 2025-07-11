from db.database_service import DatabaseService
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

class ReportService:

    @staticmethod
    def group_projects_by_importance(output_file="importance_and_type_grouped_data.xlsx"):
        """
        Group projects by importance and generate a report.
        """
        db_service = DatabaseService()
        grouped_data = db_service.fetch_grouped_project_data()
        df = pd.DataFrame(grouped_data)
        # Ensure type_description is correctly used in the pivot table and data
        df['type_description'] = df['type_description'].fillna('Unknown')
        df_pivot = df.pivot_table(index=['importance_description', 'type_description', 'branch', 'operations', 'project_description'],
                                  columns='year',
                                  values='total_investment',
                                  aggfunc='sum').reset_index()
        df_pivot.columns = [int(col) if isinstance(col, str) and col.isdigit() else col for col in df_pivot.columns]
        df_pivot['Row Total'] = df_pivot.iloc[:, 5:].sum(axis=1)
        df_pivot.loc['Column Total'] = df_pivot.iloc[:, 5:].sum(axis=0)
        df_pivot.loc['Column Total', 'Row Total'] = df_pivot['Row Total'].sum()
        
        # Merge cells for rows with the same importance
        df_pivot['importance_description'] = df_pivot['importance_description'].where(df_pivot['importance_description'] != df_pivot['importance_description'].shift())
        
        # Ensure type_description is used instead of type_id
        df_pivot['type_description'] = df_pivot['type_description'].where(df_pivot['type_description'] != df_pivot['type_description'].shift())

        # Merge cells for rows with the same branch
        df_pivot['branch'] = df_pivot['branch'].where(df_pivot['branch'] != df_pivot['branch'].shift())

        # Merge cells for rows with the same operations
        df_pivot['operations'] = df_pivot['operations'].where(df_pivot['operations'] != df_pivot['operations'].shift())
        
        # Ensure year columns are integers in Excel
        for col in df_pivot.columns[5:-1]:
            if isinstance(col, int):
                df_pivot[col] = pd.to_numeric(df_pivot[col], errors='coerce')
                
        # Create a sheet for the original data showing all importance groups
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_pivot.to_excel(writer, index=False, sheet_name='All Importance Groups')

        # Create individual sheets for each importance group
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for importance_group, importance_data in df_pivot.groupby('importance_description'):
                # Ensure all rows for the importance group are included
                importance_data = df[df['importance_description'] == importance_group].reset_index(drop=True)
                # Drop the 'importance' and 'type' columns for individual sheets
                importance_data = importance_data.drop(columns=['importance', 'type', 'importance_description'], errors='ignore')
                # Pivot the data to ensure years 2025 to 2035 are columns and investments are placed in the correct year column
                year_columns = list(range(2025, 2036))
                importance_data = importance_data.pivot_table(index=['type_description', 'branch', 'operations', 'project_description'],
                                                             columns='year',
                                                             values='total_investment',
                                                             aggfunc='sum').reset_index()
                # Ensure all year columns are present
                for year in year_columns:
                    if year not in importance_data.columns:
                        importance_data[year] = 0
                # Reorder columns to place year columns after the index columns
                importance_data = importance_data[['type_description', 'branch', 'operations', 'project_description'] + year_columns]
                # Reorder columns to place 'type_description' as the first column
                columns_order = ['type_description'] + [col for col in importance_data.columns if col != 'type_description']
                importance_data = importance_data[columns_order]
                # Merge cells for rows with the same type_description, branch, and operations
                importance_data['type_description'] = importance_data['type_description'].where(importance_data['type_description'] != importance_data['type_description'].shift())
                importance_data['branch'] = importance_data['branch'].where(importance_data['branch'] != importance_data['branch'].shift())
                importance_data['operations'] = importance_data['operations'].where(importance_data['operations'] != importance_data['operations'].shift())
                # Sanitize sheet name to remove invalid characters
                sanitized_sheet_name = ''.join(c if c.isalnum() or c.isspace() else '_' for c in str(importance_group))
                sanitized_sheet_name = sanitized_sheet_name[:31]
                # Write the entire importance group data to a single sheet
                importance_data.to_excel(writer, index=False, sheet_name=sanitized_sheet_name.strip())

        # Apply alternating row colors (blue and white) to each sheet
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            workbook = writer.book
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                blue_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), start=1):
                    fill = blue_fill if row_idx % 2 == 0 else white_fill
                    for cell in row:
                        cell.fill = fill

        print(f"[INFO] Grouped data with years as columns has been saved to {output_file}.")

    @staticmethod
    def save_dataframe_to_excel(df, output_file, sheet_name):
        """
        General method to save a DataFrame to an Excel file.
        """
        import openpyxl
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name)
        print(f"[INFO] DataFrame saved to {output_file} (sheet: {sheet_name}).")

    @staticmethod
    def create_depreciations_by_cost_center_report():
        """
        Generate a depreciation report grouped by cost center using the report repository.
        Returns the resulting DataFrame.
        """
        from db.repository_factory import RepositoryFactory
        report_repo = RepositoryFactory.create_report_repository()
        # Fetch depreciations grouped by cost center, year, and month
        data = report_repo.fetch_depreciations_by_cost_center()
        import pandas as pd
        df = pd.DataFrame(data)
        # Pivot table: cost_center as index, year and month as columns
        if not df.empty:
            df_pivot = df.pivot_table(index=['cost_center'], columns=['year', 'month'], values='total_depreciation', aggfunc='sum', fill_value=0)
        else:
            df_pivot = pd.DataFrame()
        return df_pivot

    @staticmethod
    def create_investments_by_year_report(output_file="investments_by_year.xlsx"):
        """
        Generate an investments report grouped by year.
        """
        db_service = DatabaseService()
        query = '''
            SELECT year, SUM(investment_amount) AS total_investment
            FROM investments
            GROUP BY year
            ORDER BY year
        '''
        data = db_service.execute_query(query, fetch=True)
        df = pd.DataFrame(data)
        # Write to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Investments by Year')
        print(f"[INFO] Investments by year report saved to {output_file}.")
